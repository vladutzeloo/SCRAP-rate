import os
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime, timedelta
import webbrowser
import json
from collections import defaultdict
import re
import warnings

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def main():
    """Main function - SCRAP RATE BI Dashboard Generator"""
    print("🏭 SCRAP RATE Professional BI Dashboard Generator")
    print("=" * 70)

    # Create a root window (but hide it)
    root = tk.Tk()
    root.withdraw()

    try:
        # Step 1: Select CONTROL.xlsx file
        print("📁 Select your CONTROL.xlsx file...")
        excel_file = filedialog.askopenfilename(
            title="Select CONTROL.xlsx File",
            initialdir=os.path.expanduser("~"),
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if not excel_file:
            print("❌ No file selected. Exiting...")
            return

        print(f"✅ Selected file: {excel_file}")

        # Step 2: Extract and analyze data
        print("📊 Extracting scrap rate data from Excel file...")
        scrap_data = extract_scrap_data_from_excel(excel_file)

        if not scrap_data:
            messagebox.showerror("No Data Found",
                               "No scrap data could be extracted from the Excel file.")
            return

        print(f"✅ Extracted data from {len(scrap_data['all_records'])} records")

        # Step 3: Generate BI dashboard
        html_content = generate_scrap_dashboard(scrap_data, excel_file)

        # Step 4: Save to desktop with timestamp
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        output_file = os.path.join(desktop, f"SCRAP_RATE_Dashboard_{timestamp}.html")

        print("\n" + "=" * 70)
        print("💾 Saving dashboard...")

        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)

        print(f"✅ SUCCESS! Dashboard created!")
        print("=" * 70)
        print(f"📁 LOCATION: {output_file}")
        print(f"📂 FOLDER:   {desktop}")
        print(f"📄 FILENAME: SCRAP_RATE_Dashboard_{timestamp}.html")
        print("=" * 70 + "\n")

        # Step 5: Success message and open
        show_success(output_file, scrap_data)

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        root.destroy()

def extract_scrap_data_from_excel(excel_file):
    """Extract scrap rate data from CONTROL.xlsx using openpyxl"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        messagebox.showerror("Missing Library",
                           "openpyxl is required. Install it with: pip install openpyxl")
        return None

    print("  📖 Opening Excel file...")
    wb = openpyxl.load_workbook(excel_file, data_only=True)

    data = {
        'all_records': [],
        'by_date': defaultdict(list),
        'by_machine': defaultdict(list),
        'by_controlor': defaultdict(list),
        'by_part_number': defaultdict(list),
        'sheet_names': wb.sheetnames
    }

    # Process each sheet
    for sheet_name in wb.sheetnames:
        print(f"  📄 Processing sheet: {sheet_name}")

        if sheet_name == "Drop Down List":
            continue  # Skip reference data sheet

        sheet = wb[sheet_name]

        # Find header row - special handling for "Rebuturi" sheet
        if "Rebuturi" in sheet_name:
            header_row = 2  # Headers are in row 2 for this sheet
        else:
            header_row = 1

        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(header_row, col).value
            if cell_value:
                headers.append(str(cell_value).strip())
            else:
                headers.append(f"Column_{col}")

        print(f"    📋 Headers: {headers[:10]}...")  # Print first 10 headers

        # Extract data rows
        row_count = 0
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row_data = {}
            has_data = False

            for col_idx, header in enumerate(headers, start=1):
                cell_value = sheet.cell(row_idx, col_idx).value
                row_data[header] = cell_value
                if cell_value is not None and str(cell_value).strip():
                    has_data = True

            if has_data:
                # Add sheet name to identify source
                row_data['_sheet'] = sheet_name

                # Extract key fields with flexible column name matching
                date_val = extract_field(row_data, ['Data', 'Date', 'Data/Date'])
                machine_val = extract_field(row_data, ['Machine', 'Masina'])
                controlor_val = extract_field(row_data, ['Controlor', 'Inspector'])

                # Extract quantity fields - CORRECTED COLUMN NAMES
                total_parts = extract_number(extract_field(row_data, ['Total piese\nTotal parts', 'Total piese', 'Total parts']))
                total_ok = extract_number(extract_field(row_data, ['Total piese OK\nTotal parts OK', 'Total piese OK', 'Total parts OK']))
                piese_nok = extract_number(extract_field(row_data, ['Piese NOK\n(Scrap/rework)', 'Piese NOK']))
                scrap_rebut = extract_number(extract_field(row_data, ['SCRAP\nREBUT', 'REBUT\nSCRAP', 'SCRAP', 'REBUT', 'Total piese rebut trimise\nTotal scrap parts sent']))
                quarantine = extract_number(extract_field(row_data, ['QUARANTINE\nCARANTINA\nSUSPECTE', 'CARANTINA\nQUARANTINE\nSUSPECTE', 'QUARANTINE', 'SUSPECTE', 'Quarantine']))
                derogation = extract_number(extract_field(row_data, ['DEROGATION\nDEROGARE', 'DEROGARE\nDEROGATION', 'DEROGATION']))

                # Parse date
                parsed_date = parse_date(date_val)
                row_data['_parsed_date'] = parsed_date

                # Store extracted values
                row_data['_total_parts'] = total_parts
                row_data['_total_ok'] = total_ok
                row_data['_piese_nok'] = piese_nok
                row_data['_scrap_rebut'] = scrap_rebut
                row_data['_quarantine'] = quarantine
                row_data['_derogation'] = derogation

                # Calculate total NOK (piese_nok or calculate from total - ok)
                if piese_nok is not None:
                    total_nok = piese_nok
                elif total_parts and total_ok:
                    total_nok = total_parts - total_ok
                else:
                    total_nok = 0

                row_data['_total_nok'] = total_nok

                # Calculate scrap rate based on total NOK
                if total_parts and total_parts > 0:
                    scrap_rate = (total_nok / total_parts * 100)
                    row_data['_scrap_rate'] = round(scrap_rate, 2)
                else:
                    row_data['_scrap_rate'] = None

                # Extract part numbers from the row
                part_numbers = extract_part_numbers_from_row(row_data)
                row_data['_part_numbers'] = part_numbers

                data['all_records'].append(row_data)
                row_count += 1

                # Index by date
                if parsed_date:
                    data['by_date'][parsed_date].append(row_data)

                # Index by machine
                if machine_val:
                    data['by_machine'][str(machine_val)].append(row_data)

                # Index by controlor
                if controlor_val:
                    data['by_controlor'][str(controlor_val)].append(row_data)

                # Index by part numbers
                for part_num in part_numbers:
                    data['by_part_number'][part_num].append(row_data)

        print(f"    ✓ Extracted {row_count} records from {sheet_name}")

    wb.close()

    print(f"  ✅ Total records extracted: {len(data['all_records'])}")
    print(f"  📅 Date range: {len(data['by_date'])} unique dates")
    print(f"  🏭 Machines: {len(data['by_machine'])} unique")
    print(f"  👤 Controlers: {len(data['by_controlor'])} unique")
    print(f"  🔧 Part numbers: {len(data['by_part_number'])} unique")

    return data

def extract_field(row_data, possible_names):
    """Extract field value trying multiple possible column names"""
    for name in possible_names:
        if name in row_data and row_data[name] is not None:
            return row_data[name]
    return None

def extract_number(value):
    """Extract numeric value from cell"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Try to extract number from string
        match = re.search(r'[\d,]+\.?\d*', value.replace(',', ''))
        if match:
            try:
                return float(match.group())
            except:
                return None
    return None

def parse_date(date_val):
    """Parse date from various formats"""
    if date_val is None:
        return None

    # If it's already a datetime object
    if isinstance(date_val, datetime):
        return date_val.strftime('%Y-%m-%d')

    # If it's a string
    if isinstance(date_val, str):
        date_val = date_val.strip()
        # Try various date formats
        for fmt in ['%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d', '%d-%m-%Y']:
            try:
                dt = datetime.strptime(date_val, fmt)
                return dt.strftime('%Y-%m-%d')
            except:
                continue

    return None

def extract_part_numbers_from_row(row_data):
    """Extract part numbers from row (looking for patterns like R900305231, F-688038.02-0411.WH.WE36, etc.)"""
    part_numbers = []
    part_pattern = re.compile(r'[A-Z]\d{9}|[A-Z]-\d{6}\.\d{2}-\d{4}\.[A-Z]{2}\.[A-Z]{2}\d{0,2}|\d{4}-\d{4}-\d{2}')

    for key, value in row_data.items():
        if value and isinstance(value, str):
            matches = part_pattern.findall(value)
            part_numbers.extend(matches)

    return list(set(part_numbers))  # Remove duplicates

def generate_scrap_dashboard(scrap_data, excel_file):
    """Generate professional scrap rate BI dashboard with red theme matching monthly dashboard"""

    # Calculate overall statistics
    total_records = len(scrap_data['all_records'])

    # Calculate total quantities and scrap
    total_parts = 0
    total_ok = 0
    total_nok = 0
    records_with_data = []

    for record in scrap_data['all_records']:
        parts = record.get('_total_parts', 0)
        ok = record.get('_total_ok', 0)
        nok = record.get('_total_nok', 0)

        if parts:
            total_parts += parts
        if ok:
            total_ok += ok
        if nok:
            total_nok += nok

        if record['_scrap_rate'] is not None:
            records_with_data.append(record)

    # Calculate overall scrap rate
    overall_scrap_rate = (total_nok / total_parts * 100) if total_parts > 0 else 0

    # Calculate quality rate (inverse of scrap rate)
    overall_quality_rate = (total_ok / total_parts * 100) if total_parts > 0 else 0

    # Get date range
    dates = sorted([d for d in scrap_data['by_date'].keys() if d])
    date_range = f"{dates[0]} to {dates[-1]}" if dates else "Unknown"

    # Calculate weekly and monthly scrap rates
    weekly_stats = calculate_weekly_stats(scrap_data)
    monthly_stats = calculate_monthly_stats(scrap_data)

    # Calculate daily scrap rate for last 14 days
    daily_stats = calculate_daily_stats(scrap_data, days=14)

    # Calculate statistics by machine
    machine_stats = calculate_machine_stats(scrap_data)

    # Calculate statistics by controlor
    controlor_stats = calculate_controlor_stats(scrap_data)

    # Calculate statistics by part number
    part_stats = calculate_part_stats(scrap_data)

    # Calculate trend data (scrap rate over time)
    trend_data = calculate_trend_data(scrap_data)

    # Calculate scrap category breakdown (by sheet)
    category_data = calculate_category_breakdown(scrap_data)

    # Prepare data for charts
    trend_json = json.dumps(trend_data)
    machine_stats_json = json.dumps(machine_stats)
    controlor_stats_json = json.dumps(controlor_stats)
    part_stats_json = json.dumps(part_stats)
    category_json = json.dumps(category_data)
    weekly_stats_json = json.dumps(weekly_stats)
    monthly_stats_json = json.dumps(monthly_stats)
    daily_stats_json = json.dumps(daily_stats)

    # Current date and time
    current_datetime = datetime.now().strftime("%B %d, %Y at %I:%M %p")

    # Build HTML (copying exact design from monthly dashboard)
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SCRAP RATE BI Dashboard</title>
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
    <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/3.9.1/chart.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.2.0/dist/chartjs-plugin-datalabels.min.js"></script>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: #f8f9fa;
            min-height: 100vh;
            padding: 20px;
            color: #2c3e50;
        }}

        .container {{
            max-width: 1600px;
            margin: 0 auto;
            background: #ffffff;
            border-radius: 12px;
            box-shadow: 0 4px 16px rgba(0, 0, 0, 0.1);
            overflow: hidden;
            border: 1px solid #e9ecef;
        }}

        /* Red-themed Header */
        .header {{
            background: linear-gradient(135deg, #dc2626 0%, #b91c1c 100%);
            color: white;
            padding: 30px 40px;
            border-bottom: 4px solid #ef4444;
        }}

        .header-content {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
            gap: 20px;
        }}

        .header-left {{
            display: flex;
            align-items: center;
            gap: 20px;
        }}

        .company-logo {{
            background: rgba(255, 255, 255, 0.15);
            padding: 12px 24px;
            border-radius: 8px;
            font-size: 1.5rem;
            font-weight: bold;
            letter-spacing: 2px;
            border: 1px solid rgba(255, 255, 255, 0.2);
        }}

        .header-title {{
            font-size: 2.2rem;
            font-weight: 600;
            margin: 0;
        }}

        .header-right {{
            text-align: right;
            font-size: 0.95rem;
            opacity: 0.9;
        }}

        .last-updated {{
            margin-bottom: 5px;
        }}

        .report-period {{
            font-weight: 600;
        }}

        /* Red-themed KPI Cards Grid (3 cards) */
        .kpi-grid {{
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 25px;
            padding: 40px;
            background: #f8f9fa;
        }}

        .kpi-card {{
            background: #ffffff;
            border-radius: 12px;
            padding: 30px;
            text-align: center;
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08);
            transition: all 0.3s ease;
            border-left: 4px solid transparent;
        }}

        .kpi-card:hover {{
            transform: translateY(-4px);
            box-shadow: 0 8px 25px rgba(0, 0, 0, 0.15);
        }}

        .kpi-card.scrap {{ border-left-color: #dc2626; }}
        .kpi-card.quality {{ border-left-color: #dc2626; }}
        .kpi-card.volume {{ border-left-color: #dc2626; }}

        .kpi-icon {{
            font-size: 2.5rem;
            margin-bottom: 15px;
            color: #dc2626;
        }}

        .kpi-value {{
            font-size: 2.8rem;
            font-weight: 700;
            margin-bottom: 8px;
            color: #2c3e50;
        }}

        .kpi-label {{
            color: #7f8c8d;
            font-size: 1rem;
            font-weight: 500;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        .kpi-period {{
            color: #95a5a6;
            font-size: 0.85rem;
            margin-top: 5px;
            font-style: italic;
        }}

        /* Analytics Section */
        .analytics-section {{
            padding: 40px;
            background: #ffffff;
        }}

        .section-title {{
            font-size: 1.8rem;
            color: #2c3e50;
            margin-bottom: 30px;
            padding-bottom: 10px;
            border-bottom: 2px solid #dc2626;
            font-weight: 600;
        }}

        .charts-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 30px;
            margin-bottom: 40px;
        }}

        .charts-full {{
            display: grid;
            grid-template-columns: 1fr;
            gap: 30px;
            margin-bottom: 40px;
        }}

        .chart-container {{
            background: #ffffff;
            border-radius: 12px;
            padding: 25px;
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08);
            border: 1px solid #e9ecef;
        }}

        .chart-title {{
            font-size: 1.2rem;
            color: #2c3e50;
            margin-bottom: 20px;
            text-align: center;
            font-weight: 600;
        }}

        .chart-wrapper {{
            position: relative;
            height: 400px;
        }}

        .chart-wrapper-small {{
            position: relative;
            height: 350px;
        }}

        /* Table styles */
        .table-section {{
            padding: 40px;
            background: #f8f9fa;
            border-top: 1px solid #e9ecef;
        }}

        .data-table-container {{
            background: #ffffff;
            border-radius: 12px;
            overflow: hidden;
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08);
            border: 1px solid #e9ecef;
            margin-bottom: 30px;
        }}

        .data-table {{
            width: 100%;
            border-collapse: collapse;
        }}

        .data-table th {{
            background: linear-gradient(135deg, #dc2626 0%, #b91c1c 100%);
            color: white;
            padding: 15px 12px;
            text-align: left;
            font-weight: 600;
            font-size: 0.9rem;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        .data-table td {{
            padding: 12px;
            border-bottom: 1px solid rgba(220, 38, 38, 0.1);
            font-size: 0.9rem;
            color: #2c3e50;
        }}

        .data-table tbody tr:hover {{
            background: rgba(220, 38, 38, 0.05);
        }}

        .data-table tbody tr:nth-child(even) {{
            background: #f8f9fa;
        }}

        .scrap-rate-high {{
            background: rgba(220, 38, 38, 0.15);
            color: #dc2626;
            font-weight: 600;
            padding: 4px 8px;
            border-radius: 4px;
        }}

        .scrap-rate-medium {{
            background: rgba(245, 158, 11, 0.15);
            color: #f59e0b;
            font-weight: 600;
            padding: 4px 8px;
            border-radius: 4px;
        }}

        .scrap-rate-low {{
            background: rgba(34, 197, 94, 0.15);
            color: #22c55e;
            font-weight: 600;
            padding: 4px 8px;
            border-radius: 4px;
        }}

        .footer {{
            background: #2c3e50;
            color: white;
            padding: 20px 40px;
            text-align: center;
            font-size: 0.9rem;
        }}
    </style>
</head>
<body>
    <div class="container">
        <!-- Header -->
        <div class="header">
            <div class="header-content">
                <div class="header-left">
                    <div class="company-logo">
                        <i class="fas fa-industry"></i> SCRAP
                    </div>
                    <h1 class="header-title">Scrap Rate Analytics Dashboard</h1>
                </div>
                <div class="header-right">
                    <div class="last-updated">Last Updated: {current_datetime}</div>
                    <div class="report-period">Data Period: {date_range}</div>
                    <div class="report-period">Total Records: {total_records:,}</div>
                </div>
            </div>
        </div>

        <!-- KPI Cards -->
        <div class="kpi-grid">
            <div class="kpi-card scrap">
                <div class="kpi-icon">
                    <i class="fas fa-exclamation-triangle"></i>
                </div>
                <div class="kpi-value">{overall_scrap_rate:.2f}%</div>
                <div class="kpi-label">Overall Scrap Rate</div>
                <div class="kpi-period">{total_nok:,.0f} NOK / {total_parts:,.0f} total</div>
            </div>

            <div class="kpi-card quality">
                <div class="kpi-icon">
                    <i class="fas fa-check-circle"></i>
                </div>
                <div class="kpi-value">{overall_quality_rate:.2f}%</div>
                <div class="kpi-label">Quality Rate</div>
                <div class="kpi-period">{total_ok:,.0f} OK parts</div>
            </div>

            <div class="kpi-card volume">
                <div class="kpi-icon">
                    <i class="fas fa-boxes"></i>
                </div>
                <div class="kpi-value">{total_parts:,.0f}</div>
                <div class="kpi-label">Total Parts Produced</div>
                <div class="kpi-period">Across {len(scrap_data['by_machine'])} machines</div>
            </div>
        </div>

        <!-- Analytics Section -->
        <div class="analytics-section">
            <h2 class="section-title">Scrap Rate Trends & Analysis</h2>

            <!-- Charts Grid -->
            <div class="charts-full">
                <div class="chart-container">
                    <h3 class="chart-title">📊 Weekly Scrap Rate Trend (Easier to Read)</h3>
                    <div class="chart-wrapper">
                        <canvas id="weeklyTrendChart"></canvas>
                    </div>
                </div>
            </div>

            <!-- Daily Scrap Rate Section -->
            <div class="charts-full">
                <div class="chart-container">
                    <h3 class="chart-title">📅 Daily Scrap Rate - Last 14 Days</h3>
                    <div class="chart-wrapper">
                        <canvas id="dailyTrendChart"></canvas>
                    </div>
                </div>
            </div>

            <div class="charts-grid">
                <div class="chart-container">
                    <h3 class="chart-title">Scrap Rate by Machine</h3>
                    <div class="chart-wrapper">
                        <canvas id="machineChart"></canvas>
                    </div>
                </div>

                <div class="chart-container">
                    <h3 class="chart-title">OK vs NOK Parts Distribution</h3>
                    <div class="chart-wrapper">
                        <canvas id="distributionChart"></canvas>
                    </div>
                </div>
            </div>

            <div class="charts-grid">
                <div class="chart-container">
                    <h3 class="chart-title">Scrap by Category/Sheet</h3>
                    <div class="chart-wrapper-small">
                        <canvas id="categoryChart"></canvas>
                    </div>
                </div>

                <div class="chart-container">
                    <h3 class="chart-title">Scrap Rate by Inspector</h3>
                    <div class="chart-wrapper-small">
                        <canvas id="controlorChart"></canvas>
                    </div>
                </div>
            </div>
        </div>

        <!-- Weekly & Monthly Section -->
        <div class="analytics-section">
            <h2 class="section-title">Weekly & Monthly Scrap Rate Analysis</h2>

            <div class="charts-grid">
                <!-- Weekly Scrap Rate Table -->
                <div class="data-table-container">
                    <h3 class="chart-title"><i class="fas fa-calendar-week"></i> Weekly Scrap Rate</h3>
                    <table class="data-table">
                        <thead>
                            <tr>
                                <th>Week</th>
                                <th>Parts</th>
                                <th>NOK</th>
                                <th>Scrap Rate</th>
                                <th>Quality</th>
                            </tr>
                        </thead>
                        <tbody id="weeklyTableBody">
                            <!-- Populated by JavaScript -->
                        </tbody>
                    </table>
                </div>

                <!-- Monthly Scrap Rate Table -->
                <div class="data-table-container">
                    <h3 class="chart-title"><i class="fas fa-calendar-alt"></i> Monthly Scrap Rate</h3>
                    <table class="data-table">
                        <thead>
                            <tr>
                                <th>Month</th>
                                <th>Parts</th>
                                <th>NOK</th>
                                <th>Scrap Rate</th>
                                <th>Quality</th>
                            </tr>
                        </thead>
                        <tbody id="monthlyTableBody">
                            <!-- Populated by JavaScript -->
                        </tbody>
                    </table>
                </div>
            </div>
        </div>

        <!-- Tables Section -->
        <div class="table-section">
            <h2 class="section-title">Detailed Analysis</h2>

            <!-- Machine Statistics -->
            <div class="data-table-container">
                <table class="data-table">
                    <thead>
                        <tr>
                            <th>Machine</th>
                            <th>Total Checked</th>
                            <th>Suspects/NOK</th>
                            <th>Scrap Rate</th>
                            <th>Records</th>
                        </tr>
                    </thead>
                    <tbody id="machineTableBody">
                        <!-- Populated by JavaScript -->
                    </tbody>
                </table>
            </div>

            <!-- Part Number Statistics -->
            <div class="data-table-container">
                <table class="data-table">
                    <thead>
                        <tr>
                            <th>Part Number</th>
                            <th>Total Checked</th>
                            <th>Suspects/NOK</th>
                            <th>Scrap Rate</th>
                            <th>Records</th>
                        </tr>
                    </thead>
                    <tbody id="partTableBody">
                        <!-- Populated by JavaScript -->
                    </tbody>
                </table>
            </div>
        </div>

        <!-- Footer -->
        <div class="footer">
            <p>SCRAP RATE Analytics Dashboard | Generated from {os.path.basename(excel_file)} | © {datetime.now().year}</p>
        </div>
    </div>

    <script>
        // Data from Python
        const trendData = {trend_json};
        const machineStats = {machine_stats_json};
        const controlorStats = {controlor_stats_json};
        const partStats = {part_stats_json};
        const categoryData = {category_json};
        const weeklyStats = {weekly_stats_json};
        const monthlyStats = {monthly_stats_json};
        const dailyStats = {daily_stats_json};

        // Chart colors (red theme)
        const chartColors = {{
            primary: '#dc2626',
            secondary: '#b91c1c',
            success: '#22c55e',
            danger: '#ef4444',
            warning: '#f59e0b',
            info: '#3b82f6'
        }};

        // Initialize charts on page load
        document.addEventListener('DOMContentLoaded', function() {{
            createWeeklyTrendChart();
            createDailyTrendChart();
            createMachineChart();
            createDistributionChart();
            createCategoryChart();
            createControlorChart();
            populateWeeklyMonthlyTables();
            populateTables();
        }});

        function createWeeklyTrendChart() {{
            const ctx = document.getElementById('weeklyTrendChart').getContext('2d');

            // Prepare weekly data for chart
            const weeklyLabels = weeklyStats.map(w => w.week_label);
            const weeklyRates = weeklyStats.map(w => w.scrap_rate);

            // Reverse to show oldest to newest
            weeklyLabels.reverse();
            weeklyRates.reverse();

            new Chart(ctx, {{
                type: 'line',
                data: {{
                    labels: weeklyLabels,
                    datasets: [{{
                        label: 'Weekly Scrap Rate (%)',
                        data: weeklyRates,
                        borderColor: chartColors.danger,
                        backgroundColor: 'rgba(220, 38, 38, 0.1)',
                        borderWidth: 4,
                        fill: true,
                        tension: 0.3,
                        pointRadius: 8,
                        pointHoverRadius: 10,
                        pointBackgroundColor: chartColors.danger,
                        pointBorderColor: '#ffffff',
                        pointBorderWidth: 3
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        legend: {{
                            display: true,
                            position: 'top',
                            labels: {{
                                font: {{
                                    size: 14,
                                    weight: '600'
                                }}
                            }}
                        }},
                        datalabels: {{
                            display: true,
                            align: 'top',
                            backgroundColor: 'rgba(220, 38, 38, 0.95)',
                            color: '#ffffff',
                            borderRadius: 4,
                            padding: {{
                                top: 4,
                                bottom: 4,
                                left: 8,
                                right: 8
                            }},
                            font: {{
                                size: 12,
                                weight: 'bold'
                            }},
                            formatter: function(value) {{
                                return value.toFixed(2) + '%';
                            }}
                        }},
                        tooltip: {{
                            backgroundColor: 'rgba(44, 62, 80, 0.9)',
                            titleColor: '#ffffff',
                            bodyColor: '#ffffff',
                            borderColor: chartColors.primary,
                            borderWidth: 1,
                            cornerRadius: 6,
                            padding: 12,
                            callbacks: {{
                                label: function(context) {{
                                    return 'Scrap Rate: ' + context.parsed.y.toFixed(2) + '%';
                                }}
                            }}
                        }}
                    }},
                    scales: {{
                        x: {{
                            grid: {{
                                color: 'rgba(149, 165, 166, 0.2)'
                            }},
                            ticks: {{
                                font: {{
                                    size: 11
                                }}
                            }}
                        }},
                        y: {{
                            beginAtZero: true,
                            ticks: {{
                                callback: function(value) {{
                                    return value + '%';
                                }},
                                font: {{
                                    size: 12
                                }}
                            }},
                            grid: {{
                                color: 'rgba(149, 165, 166, 0.2)'
                            }}
                        }}
                    }}
                }}
            }});
        }}

        function createDailyTrendChart() {{
            const ctx = document.getElementById('dailyTrendChart').getContext('2d');

            // Prepare daily data for chart (last 14 days)
            const dailyLabels = dailyStats.map(d => d.date_label);
            const dailyRates = dailyStats.map(d => d.scrap_rate);

            new Chart(ctx, {{
                type: 'bar',
                data: {{
                    labels: dailyLabels,
                    datasets: [{{
                        label: 'Daily Scrap Rate (%)',
                        data: dailyRates,
                        backgroundColor: dailyRates.map(rate =>
                            rate > 5 ? 'rgba(220, 38, 38, 0.8)' :
                            rate > 2 ? 'rgba(245, 158, 11, 0.8)' :
                            'rgba(34, 197, 94, 0.8)'
                        ),
                        borderColor: dailyRates.map(rate =>
                            rate > 5 ? chartColors.danger :
                            rate > 2 ? chartColors.warning :
                            chartColors.success
                        ),
                        borderWidth: 2,
                        borderRadius: 6
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        legend: {{
                            display: false
                        }},
                        datalabels: {{
                            display: true,
                            anchor: 'end',
                            align: 'top',
                            color: '#2c3e50',
                            font: {{
                                size: 11,
                                weight: 'bold'
                            }},
                            formatter: function(value) {{
                                return value.toFixed(2) + '%';
                            }}
                        }},
                        tooltip: {{
                            backgroundColor: 'rgba(44, 62, 80, 0.9)',
                            titleColor: '#ffffff',
                            bodyColor: '#ffffff',
                            borderColor: chartColors.primary,
                            borderWidth: 1,
                            cornerRadius: 6,
                            padding: 12,
                            callbacks: {{
                                label: function(context) {{
                                    const index = context.dataIndex;
                                    const day = dailyStats[index];
                                    return [
                                        'Scrap Rate: ' + day.scrap_rate.toFixed(2) + '%',
                                        'Parts: ' + day.total_parts.toLocaleString(),
                                        'NOK: ' + day.total_nok.toLocaleString()
                                    ];
                                }}
                            }}
                        }}
                    }},
                    scales: {{
                        x: {{
                            grid: {{
                                display: false
                            }},
                            ticks: {{
                                font: {{
                                    size: 10
                                }}
                            }}
                        }},
                        y: {{
                            beginAtZero: true,
                            ticks: {{
                                callback: function(value) {{
                                    return value + '%';
                                }},
                                font: {{
                                    size: 11
                                }}
                            }},
                            grid: {{
                                color: 'rgba(149, 165, 166, 0.2)'
                            }}
                        }}
                    }}
                }}
            }});
        }}

        function createMachineChart() {{
            const ctx = document.getElementById('machineChart').getContext('2d');

            // Sort and get top 10 machines
            const sortedMachines = Object.entries(machineStats)
                .sort((a, b) => b[1].scrap_rate - a[1].scrap_rate)
                .slice(0, 10);

            const labels = sortedMachines.map(m => m[0]);
            const data = sortedMachines.map(m => m[1].scrap_rate);

            new Chart(ctx, {{
                type: 'bar',
                data: {{
                    labels: labels,
                    datasets: [{{
                        label: 'Scrap Rate (%)',
                        data: data,
                        backgroundColor: chartColors.danger,
                        borderColor: chartColors.secondary,
                        borderWidth: 2,
                        borderRadius: 6,
                        hoverBackgroundColor: chartColors.secondary
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        legend: {{
                            display: false
                        }},
                        tooltip: {{
                            backgroundColor: 'rgba(44, 62, 80, 0.9)',
                            titleColor: '#ffffff',
                            bodyColor: '#ffffff',
                            borderColor: chartColors.primary,
                            borderWidth: 1,
                            cornerRadius: 6,
                            padding: 12,
                            callbacks: {{
                                label: function(context) {{
                                    const machine = sortedMachines[context.dataIndex][1];
                                    return [
                                        'Scrap Rate: ' + context.parsed.y.toFixed(2) + '%',
                                        'Total Checked: ' + machine.total_checked.toLocaleString(),
                                        'Suspects: ' + machine.total_suspecte.toLocaleString()
                                    ];
                                }}
                            }}
                        }}
                    }},
                    scales: {{
                        x: {{
                            grid: {{
                                display: false
                            }}
                        }},
                        y: {{
                            beginAtZero: true,
                            ticks: {{
                                callback: function(value) {{
                                    return value + '%';
                                }}
                            }},
                            grid: {{
                                color: 'rgba(149, 165, 166, 0.2)'
                            }}
                        }}
                    }}
                }}
            }});
        }}

        function createDistributionChart() {{
            const ctx = document.getElementById('distributionChart').getContext('2d');

            const totalOK = {total_ok};
            const totalNOK = {total_nok};

            new Chart(ctx, {{
                type: 'doughnut',
                data: {{
                    labels: ['OK Parts', 'NOK/Suspect Parts'],
                    datasets: [{{
                        data: [totalOK, totalNOK],
                        backgroundColor: [chartColors.success, chartColors.danger],
                        borderColor: [chartColors.success, chartColors.danger],
                        borderWidth: 3,
                        hoverOffset: 8
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        legend: {{
                            position: 'bottom',
                            labels: {{
                                padding: 20,
                                usePointStyle: true,
                                font: {{
                                    size: 13,
                                    weight: '600'
                                }}
                            }}
                        }},
                        tooltip: {{
                            backgroundColor: 'rgba(44, 62, 80, 0.9)',
                            titleColor: '#ffffff',
                            bodyColor: '#ffffff',
                            borderColor: chartColors.primary,
                            borderWidth: 1,
                            cornerRadius: 6,
                            padding: 12,
                            callbacks: {{
                                label: function(context) {{
                                    const value = context.raw.toLocaleString();
                                    const total = totalOK + totalNOK;
                                    const percentage = ((context.raw / total) * 100).toFixed(2);
                                    return context.label + ': ' + value + ' (' + percentage + '%)';
                                }}
                            }}
                        }}
                    }}
                }}
            }});
        }}

        function createCategoryChart() {{
            const ctx = document.getElementById('categoryChart').getContext('2d');

            const labels = categoryData.labels;
            const data = categoryData.values;

            new Chart(ctx, {{
                type: 'pie',
                data: {{
                    labels: labels,
                    datasets: [{{
                        data: data,
                        backgroundColor: [
                            chartColors.danger,
                            chartColors.warning,
                            chartColors.info,
                            chartColors.success,
                            '#8b5cf6'
                        ],
                        borderWidth: 2,
                        borderColor: '#ffffff'
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        legend: {{
                            position: 'bottom',
                            labels: {{
                                padding: 15,
                                usePointStyle: true,
                                font: {{
                                    size: 12,
                                    weight: '600'
                                }}
                            }}
                        }},
                        tooltip: {{
                            backgroundColor: 'rgba(44, 62, 80, 0.9)',
                            titleColor: '#ffffff',
                            bodyColor: '#ffffff',
                            borderColor: chartColors.primary,
                            borderWidth: 1,
                            cornerRadius: 6,
                            padding: 12
                        }}
                    }}
                }}
            }});
        }}

        function createControlorChart() {{
            const ctx = document.getElementById('controlorChart').getContext('2d');

            // Get top 10 controlers by scrap rate
            const sortedControlers = Object.entries(controlorStats)
                .sort((a, b) => b[1].scrap_rate - a[1].scrap_rate)
                .slice(0, 10);

            const labels = sortedControlers.map(c => c[0]);
            const data = sortedControlers.map(c => c[1].scrap_rate);

            new Chart(ctx, {{
                type: 'bar',
                data: {{
                    labels: labels,
                    datasets: [{{
                        label: 'Scrap Rate (%)',
                        data: data,
                        backgroundColor: chartColors.warning,
                        borderColor: '#d97706',
                        borderWidth: 2,
                        borderRadius: 6
                    }}]
                }},
                options: {{
                    indexAxis: 'y',
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        legend: {{
                            display: false
                        }},
                        tooltip: {{
                            backgroundColor: 'rgba(44, 62, 80, 0.9)',
                            titleColor: '#ffffff',
                            bodyColor: '#ffffff',
                            borderColor: chartColors.primary,
                            borderWidth: 1,
                            cornerRadius: 6,
                            padding: 12
                        }}
                    }},
                    scales: {{
                        x: {{
                            beginAtZero: true,
                            ticks: {{
                                callback: function(value) {{
                                    return value + '%';
                                }}
                            }},
                            grid: {{
                                color: 'rgba(149, 165, 166, 0.2)'
                            }}
                        }},
                        y: {{
                            grid: {{
                                display: false
                            }}
                        }}
                    }}
                }}
            }});
        }}

        function populateWeeklyMonthlyTables() {{
            // Populate weekly table
            const weeklyTableBody = document.getElementById('weeklyTableBody');
            weeklyStats.forEach((week) => {{
                const row = document.createElement('tr');
                const scrapClass = week.scrap_rate > 5 ? 'scrap-rate-high' :
                                  week.scrap_rate > 2 ? 'scrap-rate-medium' : 'scrap-rate-low';
                row.innerHTML = `
                    <td>${{week.week_label}}</td>
                    <td>${{week.total_parts.toLocaleString()}}</td>
                    <td>${{week.total_nok.toLocaleString()}}</td>
                    <td><span class="${{scrapClass}}">${{week.scrap_rate.toFixed(2)}}%</span></td>
                    <td><span class="quality-rate">${{week.quality_rate.toFixed(2)}}%</span></td>
                `;
                weeklyTableBody.appendChild(row);
            }});

            // Populate monthly table
            const monthlyTableBody = document.getElementById('monthlyTableBody');
            monthlyStats.forEach((month) => {{
                const row = document.createElement('tr');
                const scrapClass = month.scrap_rate > 5 ? 'scrap-rate-high' :
                                  month.scrap_rate > 2 ? 'scrap-rate-medium' : 'scrap-rate-low';
                row.innerHTML = `
                    <td>${{month.month_label}}</td>
                    <td>${{month.total_parts.toLocaleString()}}</td>
                    <td>${{month.total_nok.toLocaleString()}}</td>
                    <td><span class="${{scrapClass}}">${{month.scrap_rate.toFixed(2)}}%</span></td>
                    <td><span class="quality-rate">${{month.quality_rate.toFixed(2)}}%</span></td>
                `;
                monthlyTableBody.appendChild(row);
            }});
        }}

        function populateTables() {{
            // Populate machine table
            const machineTableBody = document.getElementById('machineTableBody');
            const sortedMachines = Object.entries(machineStats)
                .sort((a, b) => b[1].scrap_rate - a[1].scrap_rate);

            sortedMachines.forEach(([machine, stats]) => {{
                const row = document.createElement('tr');
                const scrapClass = stats.scrap_rate > 5 ? 'scrap-rate-high' :
                                  stats.scrap_rate > 2 ? 'scrap-rate-medium' : 'scrap-rate-low';
                row.innerHTML = `
                    <td>${{machine}}</td>
                    <td>${{stats.total_checked.toLocaleString()}}</td>
                    <td>${{stats.total_suspecte.toLocaleString()}}</td>
                    <td><span class="${{scrapClass}}">${{stats.scrap_rate.toFixed(2)}}%</span></td>
                    <td>${{stats.record_count}}</td>
                `;
                machineTableBody.appendChild(row);
            }});

            // Populate part table
            const partTableBody = document.getElementById('partTableBody');
            const sortedParts = Object.entries(partStats)
                .sort((a, b) => b[1].scrap_rate - a[1].scrap_rate)
                .slice(0, 50);  // Top 50 parts

            sortedParts.forEach(([part, stats]) => {{
                const row = document.createElement('tr');
                const scrapClass = stats.scrap_rate > 5 ? 'scrap-rate-high' :
                                  stats.scrap_rate > 2 ? 'scrap-rate-medium' : 'scrap-rate-low';
                row.innerHTML = `
                    <td>${{part}}</td>
                    <td>${{stats.total_checked.toLocaleString()}}</td>
                    <td>${{stats.total_suspecte.toLocaleString()}}</td>
                    <td><span class="${{scrapClass}}">${{stats.scrap_rate.toFixed(2)}}%</span></td>
                    <td>${{stats.record_count}}</td>
                `;
                partTableBody.appendChild(row);
            }});
        }}
    </script>
</body>
</html>"""

    return html

def calculate_weekly_stats(scrap_data):
    """Calculate scrap rate by week"""
    from datetime import datetime, timedelta

    weekly_data = {}

    for date_str, records in scrap_data['by_date'].items():
        try:
            # Parse date and get ISO week number
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            year_week = date_obj.strftime('%Y-W%U')  # Format: YYYY-W##

            if year_week not in weekly_data:
                weekly_data[year_week] = {
                    'total_parts': 0,
                    'total_ok': 0,
                    'total_nok': 0,
                    'start_date': date_obj,
                    'end_date': date_obj
                }

            # Update date range
            if date_obj < weekly_data[year_week]['start_date']:
                weekly_data[year_week]['start_date'] = date_obj
            if date_obj > weekly_data[year_week]['end_date']:
                weekly_data[year_week]['end_date'] = date_obj

            # Aggregate parts
            for record in records:
                weekly_data[year_week]['total_parts'] += record.get('_total_parts', 0) or 0
                weekly_data[year_week]['total_ok'] += record.get('_total_ok', 0) or 0
                weekly_data[year_week]['total_nok'] += record.get('_total_nok', 0) or 0
        except:
            continue

    # Calculate scrap rates and format output
    weekly_summary = []
    for week, data in sorted(weekly_data.items(), reverse=True):
        scrap_rate = (data['total_nok'] / data['total_parts'] * 100) if data['total_parts'] > 0 else 0
        quality_rate = (data['total_ok'] / data['total_parts'] * 100) if data['total_parts'] > 0 else 0

        weekly_summary.append({
            'week': week,
            'week_label': f"Week {week.split('-W')[1]} ({data['start_date'].strftime('%b %d')} - {data['end_date'].strftime('%b %d')})",
            'total_parts': data['total_parts'],
            'total_ok': data['total_ok'],
            'total_nok': data['total_nok'],
            'scrap_rate': round(scrap_rate, 2),
            'quality_rate': round(quality_rate, 2)
        })

    return weekly_summary

def calculate_monthly_stats(scrap_data):
    """Calculate scrap rate by month"""
    from datetime import datetime

    monthly_data = {}

    for date_str, records in scrap_data['by_date'].items():
        try:
            # Parse date and get year-month
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            year_month = date_obj.strftime('%Y-%m')  # Format: YYYY-MM

            if year_month not in monthly_data:
                monthly_data[year_month] = {
                    'total_parts': 0,
                    'total_ok': 0,
                    'total_nok': 0
                }

            # Aggregate parts
            for record in records:
                monthly_data[year_month]['total_parts'] += record.get('_total_parts', 0) or 0
                monthly_data[year_month]['total_ok'] += record.get('_total_ok', 0) or 0
                monthly_data[year_month]['total_nok'] += record.get('_total_nok', 0) or 0
        except:
            continue

    # Calculate scrap rates and format output
    monthly_summary = []
    for month, data in sorted(monthly_data.items(), reverse=True):
        scrap_rate = (data['total_nok'] / data['total_parts'] * 100) if data['total_parts'] > 0 else 0
        quality_rate = (data['total_ok'] / data['total_parts'] * 100) if data['total_parts'] > 0 else 0

        # Parse month for display
        try:
            month_obj = datetime.strptime(month, '%Y-%m')
            month_label = month_obj.strftime('%B %Y')
        except:
            month_label = month

        monthly_summary.append({
            'month': month,
            'month_label': month_label,
            'total_parts': data['total_parts'],
            'total_ok': data['total_ok'],
            'total_nok': data['total_nok'],
            'scrap_rate': round(scrap_rate, 2),
            'quality_rate': round(quality_rate, 2)
        })

    return monthly_summary

def calculate_daily_stats(scrap_data, days=14):
    """Calculate daily scrap rate for last N days"""
    from datetime import datetime, timedelta

    # Get the latest date in data
    sorted_dates = sorted(scrap_data['by_date'].keys(), reverse=True)
    if not sorted_dates:
        return []

    try:
        latest_date = datetime.strptime(sorted_dates[0], '%Y-%m-%d')
    except:
        return []

    # Calculate cutoff date
    cutoff_date = latest_date - timedelta(days=days-1)

    daily_summary = []

    for date_str in sorted_dates[:days]:  # Get last N days
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')

            if date_obj < cutoff_date:
                continue

            records = scrap_data['by_date'][date_str]
            total_parts = 0
            total_ok = 0
            total_nok = 0

            for record in records:
                total_parts += record.get('_total_parts', 0) or 0
                total_ok += record.get('_total_ok', 0) or 0
                total_nok += record.get('_total_nok', 0) or 0

            if total_parts > 0:
                scrap_rate = (total_nok / total_parts * 100)
                quality_rate = (total_ok / total_parts * 100)

                daily_summary.append({
                    'date': date_str,
                    'date_label': date_obj.strftime('%b %d'),
                    'total_parts': total_parts,
                    'total_ok': total_ok,
                    'total_nok': total_nok,
                    'scrap_rate': round(scrap_rate, 2),
                    'quality_rate': round(quality_rate, 2)
                })
        except:
            continue

    # Reverse to show oldest to newest
    daily_summary.reverse()

    return daily_summary

def calculate_machine_stats(scrap_data):
    """Calculate statistics by machine"""
    machine_stats = {}

    for machine, records in scrap_data['by_machine'].items():
        total_parts = 0
        total_ok = 0
        total_nok = 0

        for record in records:
            parts = record.get('_total_parts', 0)
            ok = record.get('_total_ok', 0)
            nok = record.get('_total_nok', 0)

            if parts:
                total_parts += parts
            if ok:
                total_ok += ok
            if nok:
                total_nok += nok

        scrap_rate = (total_nok / total_parts * 100) if total_parts > 0 else 0

        machine_stats[machine] = {
            'total_checked': total_parts,
            'total_suspecte': total_nok,
            'scrap_rate': round(scrap_rate, 2),
            'record_count': len(records)
        }

    return machine_stats

def calculate_controlor_stats(scrap_data):
    """Calculate statistics by controlor/inspector"""
    controlor_stats = {}

    for controlor, records in scrap_data['by_controlor'].items():
        total_parts = 0
        total_ok = 0
        total_nok = 0

        for record in records:
            parts = record.get('_total_parts', 0)
            ok = record.get('_total_ok', 0)
            nok = record.get('_total_nok', 0)

            if parts:
                total_parts += parts
            if ok:
                total_ok += ok
            if nok:
                total_nok += nok

        scrap_rate = (total_nok / total_parts * 100) if total_parts > 0 else 0

        controlor_stats[controlor] = {
            'total_checked': total_parts,
            'total_suspecte': total_nok,
            'scrap_rate': round(scrap_rate, 2),
            'record_count': len(records)
        }

    return controlor_stats

def calculate_part_stats(scrap_data):
    """Calculate statistics by part number"""
    part_stats = {}

    for part, records in scrap_data['by_part_number'].items():
        total_parts = 0
        total_ok = 0
        total_nok = 0

        for record in records:
            parts = record.get('_total_parts', 0)
            ok = record.get('_total_ok', 0)
            nok = record.get('_total_nok', 0)

            if parts:
                total_parts += parts
            if ok:
                total_ok += ok
            if nok:
                total_nok += nok

        scrap_rate = (total_nok / total_parts * 100) if total_parts > 0 else 0

        part_stats[part] = {
            'total_checked': total_parts,
            'total_suspecte': total_nok,
            'scrap_rate': round(scrap_rate, 2),
            'record_count': len(records)
        }

    return part_stats

def calculate_trend_data(scrap_data):
    """Calculate scrap rate trend over time"""
    trend = {
        'labels': [],
        'scrap_rates': [],
        'volumes': []
    }

    # Sort dates
    sorted_dates = sorted(scrap_data['by_date'].keys())

    for date in sorted_dates:
        records = scrap_data['by_date'][date]
        total_parts = 0
        total_ok = 0
        total_nok = 0

        for record in records:
            parts = record.get('_total_parts', 0)
            ok = record.get('_total_ok', 0)
            nok = record.get('_total_nok', 0)

            if parts:
                total_parts += parts
            if ok:
                total_ok += ok
            if nok:
                total_nok += nok

        if total_parts > 0:
            scrap_rate = (total_nok / total_parts * 100)
            trend['labels'].append(date)
            trend['scrap_rates'].append(round(scrap_rate, 2))
            trend['volumes'].append(total_parts)

    return trend

def calculate_category_breakdown(scrap_data):
    """Calculate scrap breakdown by category (sheet name)"""
    category_breakdown = {
        'labels': [],
        'values': []
    }

    sheet_stats = {}

    for record in scrap_data['all_records']:
        sheet = record.get('_sheet', 'Unknown')
        nok = record.get('_total_nok', 0)

        if nok:
            if sheet not in sheet_stats:
                sheet_stats[sheet] = 0
            sheet_stats[sheet] += nok

    # Sort by value
    sorted_sheets = sorted(sheet_stats.items(), key=lambda x: x[1], reverse=True)

    for sheet, value in sorted_sheets:
        category_breakdown['labels'].append(sheet)
        category_breakdown['values'].append(value)

    return category_breakdown

def show_success(output_file, scrap_data):
    """Show success message"""
    total_records = len(scrap_data['all_records'])

    instructions = f"""
🌟 SCRAP RATE BI DASHBOARD CREATED!

📊 DASHBOARD FEATURES:
• Total Records Analyzed: {total_records:,}
• Interactive Charts: Trend, Machine, Distribution, Category, Inspector
• Detailed Tables: Machine and Part Number Statistics
• Red Theme: Professional matching monthly dashboard design

📈 VISUALIZATIONS INCLUDED:
• Scrap Rate Trend Over Time (Line Chart)
• Scrap Rate by Machine (Bar Chart)
• OK vs NOK Parts Distribution (Doughnut Chart)
• Scrap by Category/Sheet (Pie Chart)
• Scrap Rate by Inspector (Horizontal Bar Chart)
• Machine Statistics Table
• Part Number Statistics Table

🎨 DESIGN:
• Same red-themed professional design as monthly dashboard
• Responsive and interactive charts using Chart.js
• Hover tooltips with detailed information
• Color-coded scrap rates (red=high, yellow=medium, green=low)

Perfect for analyzing scrap rate trends and identifying improvement opportunities!
    """

    desktop = os.path.dirname(output_file)
    filename = os.path.basename(output_file)

    result = messagebox.askyesno(
        "Scrap Rate Dashboard Ready!",
        f"🌟 Scrap rate dashboard created with {total_records:,} records!\n\n"
        f"📁 SAVED TO: {desktop}\n"
        f"📄 FILE: {filename}\n\n"
        f"✅ Interactive trend analysis\n"
        f"✅ Machine and part number breakdown\n"
        f"✅ Inspector performance analysis\n"
        f"✅ Professional red-themed design\n"
        f"✅ Detailed statistics tables\n\n"
        f"Open dashboard now?"
    )

    if result:
        webbrowser.open(f"file://{os.path.abspath(output_file)}")
        messagebox.showinfo("Dashboard Features", instructions)

if __name__ == "__main__":
    main()
